import os
import base64
from io import BytesIO
from typing import List, Tuple, Optional

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart
import xlrd

from ..database import get_db
from ..config import MAX_FILE_SIZE, ALLOWED_EXTENSIONS
from .. import crud, schemas

router = APIRouter(prefix="/api", tags=["excel"])


def get_file_extension(filename: str) -> str:
    """获取文件扩展名"""
    return os.path.splitext(filename)[1].lower()


def detect_table_regions(rows: List[List], column_count: int) -> List[Tuple[int, int, int, int, int, int, str]]:
    """
    检测单个Sheet内的多个表格区域
    通过识别空行来分割不同的表格
    返回: List of (region_index, start_row, start_col, end_row, end_col, header_rows, table_name)
    """
    if not rows:
        return []

    regions = []
    current_start = None
    region_index = 0

    for row_idx, row in enumerate(rows):
        # 检查这一行是否全为空
        is_empty_row = all(cell is None or str(cell).strip() == "" for cell in row)

        if is_empty_row:
            # 如果当前有活跃的表格区域，结束它
            if current_start is not None:
                end_row = row_idx - 1
                if end_row >= current_start:
                    # 尝试从第一个非空单元格获取表格名称
                    table_name = None
                    first_row = rows[current_start]
                    for cell in first_row:
                        if cell is not None and str(cell).strip():
                            table_name = str(cell).strip()[:50]  # 限制长度
                            break

                    regions.append((
                        region_index,
                        current_start,
                        0,
                        end_row,
                        column_count - 1,
                        1,  # 默认1行表头
                        table_name
                    ))
                    region_index += 1
                current_start = None
        else:
            # 如果还没有开始新区域，标记开始
            if current_start is None:
                current_start = row_idx

    # 处理最后一个区域
    if current_start is not None:
        end_row = len(rows) - 1
        table_name = None
        first_row = rows[current_start]
        for cell in first_row:
            if cell is not None and str(cell).strip():
                table_name = str(cell).strip()[:50]
                break

        regions.append((
            region_index,
            current_start,
            0,
            end_row,
            column_count - 1,
            1,
            table_name
        ))

    # 如果只检测到一个区域，说明整个Sheet就是一个表格，不需要特别标记
    if len(regions) == 1:
        return []

    return regions


def get_chart_type_name(chart) -> str:
    """获取图表类型名称"""
    chart_type_map = {
        'BarChart': 'bar',
        'BarChart3D': 'bar3d',
        'LineChart': 'line',
        'LineChart3D': 'line3d',
        'PieChart': 'pie',
        'PieChart3D': 'pie3d',
        'AreaChart': 'area',
        'AreaChart3D': 'area3d',
        'ScatterChart': 'scatter',
        'RadarChart': 'radar',
        'DoughnutChart': 'doughnut',
        'BubbleChart': 'bubble',
        'StockChart': 'stock',
        'SurfaceChart': 'surface',
        'SurfaceChart3D': 'surface3d',
    }
    return chart_type_map.get(type(chart).__name__, 'unknown')


def extract_chart_data(chart) -> Optional[dict]:
    """提取图表数据"""
    try:
        chart_data = {
            'title': chart.title.text if chart.title else None,
            'series': []
        }

        for series in chart.series:
            series_info = {
                'title': str(series.title) if series.title else None,
            }
            chart_data['series'].append(series_info)

        return chart_data
    except Exception:
        return None


def parse_xlsx(file_data: bytes) -> List[dict]:
    """解析.xlsx文件，支持合并单元格、图片和图表"""
    # 使用非只读模式以获取合并单元格、图片和图表信息
    workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
    sheets_data = []

    for idx, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]

        # 获取所有行数据
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))

        # 计算实际行列数
        row_count = len(rows)
        column_count = max((len(row) for row in rows), default=0) if rows else 0

        # 收集单元格数据
        cells = []
        for row_idx, row in enumerate(rows):
            for col_idx, value in enumerate(row):
                if value is not None:
                    cells.append((row_idx, col_idx, value))

        # 收集合并单元格信息
        merged_cells = []
        for merged_range in sheet.merged_cells.ranges:
            merged_cells.append((
                merged_range.min_row - 1,  # 转换为0索引
                merged_range.min_col - 1,
                merged_range.max_row - 1,
                merged_range.max_col - 1
            ))

        # 收集图片信息
        images = []
        if hasattr(sheet, '_images'):
            for img in sheet._images:
                try:
                    # 获取图片数据
                    img_data = img._data() if callable(img._data) else img._data
                    # 获取锚定位置
                    anchor_row = 0
                    anchor_col = 0
                    if hasattr(img, 'anchor'):
                        if hasattr(img.anchor, '_from'):
                            anchor_row = img.anchor._from.row
                            anchor_col = img.anchor._from.col
                        elif hasattr(img.anchor, 'row'):
                            anchor_row = img.anchor.row
                            anchor_col = img.anchor.col

                    # 获取图片格式
                    img_format = 'png'
                    if hasattr(img, 'format'):
                        img_format = img.format.lower()
                    elif hasattr(img, 'path') and img.path:
                        img_format = os.path.splitext(img.path)[1].lstrip('.').lower() or 'png'

                    # 获取尺寸
                    width = int(img.width) if hasattr(img, 'width') and img.width else None
                    height = int(img.height) if hasattr(img, 'height') and img.height else None

                    images.append({
                        'data': img_data,
                        'format': img_format,
                        'anchor_row': anchor_row,
                        'anchor_col': anchor_col,
                        'width': width,
                        'height': height
                    })
                except Exception:
                    continue

        # 收集图表信息
        charts = []
        if hasattr(sheet, '_charts'):
            for chart in sheet._charts:
                try:
                    anchor_row = 0
                    anchor_col = 0
                    if hasattr(chart, 'anchor'):
                        if hasattr(chart.anchor, '_from'):
                            anchor_row = chart.anchor._from.row
                            anchor_col = chart.anchor._from.col

                    chart_type = get_chart_type_name(chart)
                    chart_title = chart.title.text if chart.title else None
                    chart_data = extract_chart_data(chart)

                    # 获取图表尺寸
                    width = int(chart.width * 96) if hasattr(chart, 'width') else None  # 转换为像素
                    height = int(chart.height * 96) if hasattr(chart, 'height') else None

                    charts.append({
                        'type': chart_type,
                        'title': chart_title,
                        'data': chart_data,
                        'anchor_row': anchor_row,
                        'anchor_col': anchor_col,
                        'width': width,
                        'height': height
                    })
                except Exception:
                    continue

        # 检测多表格区域
        table_regions = detect_table_regions(rows, column_count)

        sheets_data.append({
            "name": sheet_name,
            "index": idx,
            "row_count": row_count,
            "column_count": column_count,
            "cells": cells,
            "merged_cells": merged_cells,
            "images": images,
            "charts": charts,
            "table_regions": table_regions
        })

    workbook.close()
    return sheets_data


def parse_xls(file_data: bytes) -> List[dict]:
    """解析.xls文件，支持合并单元格"""
    workbook = xlrd.open_workbook(file_contents=file_data, formatting_info=True)
    sheets_data = []

    for idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(idx)

        # 收集所有行数据用于表格区域检测
        rows = []
        for row_idx in range(sheet.nrows):
            row = []
            for col_idx in range(sheet.ncols):
                value = sheet.cell_value(row_idx, col_idx)
                row.append(value if value != "" else None)
            rows.append(row)

        # 收集单元格数据
        cells = []
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                value = sheet.cell_value(row_idx, col_idx)
                if value != "":
                    cells.append((row_idx, col_idx, value))

        # 收集合并单元格信息
        merged_cells = []
        for merged_range in sheet.merged_cells:
            # xlrd的格式是 (row_start, row_end, col_start, col_end)
            # 其中end是不包含的
            row_start, row_end, col_start, col_end = merged_range
            merged_cells.append((
                row_start,
                col_start,
                row_end - 1,  # 转换为包含的索引
                col_end - 1
            ))

        # 检测多表格区域
        table_regions = detect_table_regions(rows, sheet.ncols)

        sheets_data.append({
            "name": sheet.name,
            "index": idx,
            "row_count": sheet.nrows,
            "column_count": sheet.ncols,
            "cells": cells,
            "merged_cells": merged_cells,
            "images": [],  # xls格式图片提取较复杂，暂不支持
            "charts": [],  # xls格式图表提取较复杂，暂不支持
            "table_regions": table_regions
        })

    return sheets_data


@router.post("/upload", response_model=schemas.UploadResponse)
async def upload_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """上传Excel文件"""
    # 检查文件扩展名
    ext = get_file_extension(file.filename)
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"不支持的文件格式，仅支持: {', '.join(ALLOWED_EXTENSIONS)}")

    # 读取文件内容
    file_data = await file.read()
    file_size = len(file_data)

    # 检查文件大小
    if file_size > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail=f"文件大小超出限制（最大{MAX_FILE_SIZE // 1024 // 1024}MB）")

    # 解析Excel
    try:
        if ext == ".xlsx":
            sheets_data = parse_xlsx(file_data)
        else:
            sheets_data = parse_xls(file_data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel文件解析失败: {str(e)}")

    # 保存文件信息
    db_file = crud.create_excel_file(
        db=db,
        filename=file.filename,
        file_data=file_data,
        file_size=file_size,
        sheet_count=len(sheets_data)
    )

    # 保存Sheet和数据
    for sheet_info in sheets_data:
        db_sheet = crud.create_excel_sheet(
            db=db,
            file_id=db_file.id,
            sheet_name=sheet_info["name"],
            sheet_index=sheet_info["index"],
            row_count=sheet_info["row_count"],
            column_count=sheet_info["column_count"]
        )

        # 批量保存单元格数据
        if sheet_info["cells"]:
            crud.bulk_create_excel_data(db, db_sheet.id, sheet_info["cells"])

        # 批量保存合并单元格信息
        if sheet_info.get("merged_cells"):
            crud.bulk_create_merged_cells(db, db_sheet.id, sheet_info["merged_cells"])

        # 保存图片
        for img_info in sheet_info.get("images", []):
            crud.create_sheet_image(
                db=db,
                sheet_id=db_sheet.id,
                image_data=img_info["data"],
                image_format=img_info["format"],
                anchor_row=img_info["anchor_row"],
                anchor_col=img_info["anchor_col"],
                width=img_info.get("width"),
                height=img_info.get("height")
            )

        # 保存图表
        for chart_info in sheet_info.get("charts", []):
            crud.create_sheet_chart(
                db=db,
                sheet_id=db_sheet.id,
                chart_type=chart_info["type"],
                anchor_row=chart_info["anchor_row"],
                anchor_col=chart_info["anchor_col"],
                chart_title=chart_info.get("title"),
                chart_data=chart_info.get("data"),
                width=chart_info.get("width"),
                height=chart_info.get("height")
            )

        # 保存表格区域
        if sheet_info.get("table_regions"):
            crud.bulk_create_table_regions(db, db_sheet.id, sheet_info["table_regions"])

    return schemas.UploadResponse(
        id=db_file.id,
        filename=db_file.filename,
        message="上传成功"
    )


@router.get("/files", response_model=schemas.FileListResponse)
def get_files(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db)
):
    """获取文件列表"""
    total, files = crud.get_files(db, skip=skip, limit=limit)
    return schemas.FileListResponse(
        total=total,
        items=[schemas.FileInfo.model_validate(f) for f in files]
    )


@router.get("/files/{file_id}", response_model=schemas.FileDetail)
def get_file_detail(file_id: int, db: Session = Depends(get_db)):
    """获取文件详情（包含Sheet列表）"""
    db_file = crud.get_file_by_id(db, file_id)
    if not db_file:
        raise HTTPException(status_code=404, detail="文件不存在")
    return schemas.FileDetail.model_validate(db_file)


@router.get("/files/{file_id}/sheets/{sheet_id}/data", response_model=schemas.SheetDataResponse)
def get_sheet_data(
    file_id: int,
    sheet_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=50000),
    db: Session = Depends(get_db)
):
    """获取Sheet数据（分页），包含合并单元格、图片、图表和表格区域信息"""
    # 验证文件存在
    db_file = crud.get_file_by_id(db, file_id)
    if not db_file:
        raise HTTPException(status_code=404, detail="文件不存在")

    # 验证Sheet存在且属于该文件
    db_sheet = crud.get_sheet_by_id(db, sheet_id)
    if not db_sheet or db_sheet.file_id != file_id:
        raise HTTPException(status_code=404, detail="Sheet不存在")

    # 获取分页数据
    data_records, total_rows = crud.get_sheet_data(db, sheet_id, page, page_size)

    # 将数据转换为二维数组格式
    # 首先确定列数
    column_count = db_sheet.column_count

    # 生成表头（第一行数据或列索引）
    headers = []
    if total_rows > 0:
        # 获取第一行作为表头
        header_cells = {d.column_index: d.cell_value for d in data_records if d.row_index == 0}
        if header_cells:
            headers = [header_cells.get(i, f"列{i+1}") or f"列{i+1}" for i in range(column_count)]
        else:
            headers = [f"列{i+1}" for i in range(column_count)]
    else:
        headers = [f"列{i+1}" for i in range(column_count)]

    # 组织数据为二维数组
    data_dict = {}
    for record in data_records:
        if record.row_index not in data_dict:
            data_dict[record.row_index] = {}
        data_dict[record.row_index][record.column_index] = record.cell_value

    # 转换为列表格式
    rows = []
    start_row = (page - 1) * page_size
    for row_idx in range(start_row, min(start_row + page_size, total_rows + 1)):
        if row_idx in data_dict:
            row = [data_dict[row_idx].get(col_idx, "") for col_idx in range(column_count)]
            rows.append(row)

    # 获取合并单元格信息
    merged_cells = crud.get_sheet_merged_cells(db, sheet_id)
    merged_cells_info = [
        schemas.MergedCellInfo(
            start_row=mc.start_row,
            start_col=mc.start_col,
            end_row=mc.end_row,
            end_col=mc.end_col
        )
        for mc in merged_cells
    ]

    # 获取图片信息（不包含二进制数据，只返回元信息）
    images = crud.get_sheet_images(db, sheet_id)
    images_info = [
        schemas.ImageInfo(
            id=img.id,
            image_format=img.image_format,
            anchor_row=img.anchor_row,
            anchor_col=img.anchor_col,
            width=img.width,
            height=img.height
        )
        for img in images
    ]

    # 获取图表信息
    charts = crud.get_sheet_charts(db, sheet_id)
    charts_info = [
        schemas.ChartInfo(
            id=chart.id,
            chart_type=chart.chart_type,
            chart_title=chart.chart_title,
            chart_data=chart.chart_data,
            anchor_row=chart.anchor_row,
            anchor_col=chart.anchor_col,
            width=chart.width,
            height=chart.height
        )
        for chart in charts
    ]

    # 获取表格区域信息
    table_regions = crud.get_sheet_table_regions(db, sheet_id)
    table_regions_info = [
        schemas.TableRegionInfo(
            id=tr.id,
            region_index=tr.region_index,
            start_row=tr.start_row,
            start_col=tr.start_col,
            end_row=tr.end_row,
            end_col=tr.end_col,
            header_rows=tr.header_rows,
            table_name=tr.table_name
        )
        for tr in table_regions
    ]

    return schemas.SheetDataResponse(
        sheet_id=sheet_id,
        sheet_name=db_sheet.sheet_name,
        total_rows=total_rows + 1,  # 包含表头行
        total_columns=column_count,
        page=page,
        page_size=page_size,
        headers=headers,
        data=rows,
        merged_cells=merged_cells_info,
        images=images_info,
        charts=charts_info,
        table_regions=table_regions_info
    )


@router.get("/files/{file_id}/download")
def download_file(file_id: int, db: Session = Depends(get_db)):
    """下载原始Excel文件"""
    db_file = crud.get_file_by_id(db, file_id)
    if not db_file:
        raise HTTPException(status_code=404, detail="文件不存在")

    # 确定Content-Type
    ext = get_file_extension(db_file.filename)
    if ext == ".xlsx":
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        media_type = "application/vnd.ms-excel"

    return StreamingResponse(
        BytesIO(db_file.file_data),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{db_file.filename}"'
        }
    )


@router.get("/images/{image_id}")
def get_image(image_id: int, db: Session = Depends(get_db)):
    """获取图片二进制数据"""
    db_image = crud.get_sheet_image_by_id(db, image_id)
    if not db_image:
        raise HTTPException(status_code=404, detail="图片不存在")

    # 确定Content-Type
    format_to_mime = {
        'png': 'image/png',
        'jpeg': 'image/jpeg',
        'jpg': 'image/jpeg',
        'gif': 'image/gif',
        'bmp': 'image/bmp',
        'webp': 'image/webp',
    }
    media_type = format_to_mime.get(db_image.image_format.lower(), 'application/octet-stream')

    return Response(
        content=db_image.image_data,
        media_type=media_type
    )


@router.delete("/files/{file_id}", response_model=schemas.MessageResponse)
def delete_file(file_id: int, db: Session = Depends(get_db)):
    """删除文件"""
    if not crud.delete_file(db, file_id):
        raise HTTPException(status_code=404, detail="文件不存在")
    return schemas.MessageResponse(message="删除成功")
